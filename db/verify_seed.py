#!/usr/bin/env python3
"""
Проверка собранной базы: сходится ли она с текстовыми справочниками и работают
ли запросы, ради которых схема и затевалась.

Количества пересчитываются из CSV заново, а не берутся из отчёта сборщика.

Запуск:
    python3 db/verify_seed.py [путь/к/seed.sqlite] [путь/к/Индексатор.xlsm]

Второй аргумент необязателен. Если указать исходный Excel, добавится сверка
текстовых справочников с ним — это нужно только после обновления справочников
через db/tools/extract_from_xlsm.py.
"""

import csv
import re
import sqlite3
import sys
import unicodedata
import warnings
from pathlib import Path

DB_DIR = Path(__file__).resolve().parent
SEED_DIR = DB_DIR / "seed"

ok_count = 0
fail_count = 0


def check(title, condition, detail=""):
    global ok_count, fail_count
    if condition:
        ok_count += 1
        print(f"  [ок]     {title}" + (f" — {detail}" if detail else ""))
    else:
        fail_count += 1
        print(f"  [ОШИБКА] {title}" + (f" — {detail}" if detail else ""))


def norm(value) -> str:
    if value is None:
        return ""
    s = unicodedata.normalize("NFC", str(value)).strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", s)


def read_csv(name):
    with (SEED_DIR / name).open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def _utf8_stdout() -> None:
    """Windows запускает Python с кодировкой cp1252, и любой вывод кириллицей
    роняет скрипт с UnicodeEncodeError. Переключаем потоки на UTF-8."""
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except (AttributeError, ValueError):
            pass


def main() -> int:
    _utf8_stdout()
    db_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DB_DIR / "seed.sqlite"
    xlsm_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None

    db = sqlite3.connect(db_path)
    db.row_factory = sqlite3.Row
    q = lambda sql, *a: db.execute(sql, a).fetchall()
    one = lambda sql, *a: db.execute(sql, a).fetchone()[0]

    print("\n1. Целостность базы")
    check("integrity_check", one("PRAGMA integrity_check") == "ok")
    check("foreign_key_check", len(q("PRAGMA foreign_key_check")) == 0)
    check("версия схемы записана", one("SELECT version FROM schema_version") == 4)

    print("\n2. Сверка с текстовыми справочниками")
    for csv_name, table in [("name_dict.csv", "name_dict"), ("lookup.csv", "lookup"),
                            ("lookup_kind.csv", "lookup_kind"), ("role.csv", "role"),
                            ("place.csv", "place"), ("setting.csv", "setting")]:
        n_csv = len(read_csv(csv_name))
        # к настройкам сборщик добавляет отпечаток поставки — его в CSV нет
        expected = n_csv + 1 if table == "setting" else n_csv
        n_db = one(f"SELECT count(*) FROM {table}")
        check(f"{table} перенесена полностью", expected == n_db, f"ожидалось {expected}, в базе {n_db}")

    kinds_csv = {r["kind"] for r in read_csv("lookup.csv")}
    kinds_db = {r["kind"] for r in q("SELECT DISTINCT kind FROM lookup")}
    check("состав перечней совпадает", kinds_csv == kinds_db,
          f"{len(kinds_db)} перечней")
    check("у каждого перечня есть название",
          one("SELECT count(*) FROM lookup l LEFT JOIN lookup_kind k USING(kind) WHERE k.kind IS NULL") == 0)

    print("\n3. Роли и коридоры возраста")
    check("21 роль", one("SELECT count(*) FROM role") == 21)
    for section, expected in [(1, 5), (2, 9), (3, 4), (0, 3)]:
        check(f"ролей в разделе {section}",
              one("SELECT count(*) FROM role WHERE section=?", section) == expected)
    check("коридор родителей 17-55",
          one("SELECT age_min||'-'||age_max FROM role WHERE code='father'") == "17-55")
    check("коридор восприемников 6-60",
          one("SELECT age_min||'-'||age_max FROM role WHERE code='godparent1'") == "6-60")
    check("коридор поручителей 17-60",
          one("SELECT age_min||'-'||age_max FROM role WHERE code='witness1'") == "17-60")

    print("\n4. Поиск по префиксу на кириллице")
    # Ради этого в схеме есть колонки *_norm. Проверка намеренно фиксирует, что
    # встроенный NOCASE кириллицу не понимает — чтобы схему не «упростили» обратно.
    naive = one("SELECT count(*) FROM lookup WHERE kind='rank_m' AND value LIKE 'КР%' COLLATE NOCASE")
    smart = one("SELECT count(*) FROM lookup WHERE kind='rank_m' AND value_norm LIKE 'кр%'")
    check("NOCASE на кириллице действительно не работает", naive == 0,
          f"верхний регистр через NOCASE нашёл {naive}")
    check("поиск по value_norm работает", smart > 0, f"«кр» → {smart} звани(й)")
    top = [r["value"] for r in q(
        "SELECT value FROM lookup WHERE kind='rank_m' AND value_norm LIKE 'кр%' ORDER BY sort_order LIMIT 3")]
    check("в выдаче есть «крестьянин»", "крестьянин" in top, ", ".join(top))
    n_ye = one("SELECT count(*) FROM name_dict WHERE name LIKE '%ё%'")
    check("имена с «ё» нормализованы через «е»",
          one("SELECT count(*) FROM name_dict WHERE name_norm LIKE '%ё%'") == 0,
          f"имён с «ё» в словаре: {n_ye}")

    print("\n5. Звания, которые искал Роман")
    # 13.08.2026 он искал «крестьянскую жену» в поле мужских званий и решил,
    # что звания пропали. Проверяем оба перечня и тот самый запрос.
    for value, kind in [("крестьянская жена", "rank_f"), ("крестьянская вдова", "rank_f"),
                        ("крестьянка", "rank_f"), ("крестьянин", "rank_m")]:
        found = one("SELECT count(*) FROM lookup WHERE value=? AND kind=?", value, kind)
        check(f"«{value}» есть в перечне {kind}", found == 1)
    n_f = one("SELECT count(*) FROM lookup WHERE kind='rank_f' AND value_norm LIKE 'кресть%'")
    check("по «кресть» в женских званиях что-то находится", n_f >= 3, f"{n_f} значений")

    # Запрос состава справочников — тот же, что выполняет команда lookup_summary.
    summary = q("""SELECT k.kind, k.title, count(l.id)
                     FROM lookup_kind k LEFT JOIN lookup l ON l.kind = k.kind
                    GROUP BY k.kind, k.title
                    ORDER BY count(l.id) DESC, k.title""")
    check("состав справочников считается", len(summary) == 14, f"{len(summary)} перечней")
    check("в сводке нет перечней без названия", all(r[1] for r in summary))

    print("\n6. Ранжирование подсказок по частоте")
    # Требование А-1: текущее дело → приход → вся база → словарь,
    # внутри группы по убыванию частоты.
    db.execute("INSERT INTO mk_case (id, church, village, parish_key, year) "
               "VALUES (1,'Христорождественская','Борисоглебское','test-parish',1893)")
    db.executemany(
        "INSERT INTO usage_stat (kind, scope, scope_key, value, value_norm, count) VALUES (?,?,?,?,?,?)",
        [(k, s, sk, v, norm(v), c) for k, s, sk, v, c in [
            ("rank_m", "case",   "1",           "отставной солдат",  3),
            ("rank_m", "case",   "1",           "крестьянин",        9),
            ("rank_m", "parish", "test-parish", "крестьянский сын", 40),
            ("rank_m", "global", "",            "купец",            99),
        ]])

    rank_sql = """
    WITH ranked AS (
      SELECT value,
             CASE scope WHEN 'case' THEN 1 WHEN 'parish' THEN 2 ELSE 3 END AS tier,
             count
        FROM usage_stat
       WHERE kind = ? AND value_norm LIKE ?
         AND ((scope='case'   AND scope_key = ?)
           OR (scope='parish' AND scope_key = ?)
           OR  scope='global')
      UNION ALL
      SELECT value, 4 AS tier, 0 FROM lookup WHERE kind = ? AND value_norm LIKE ?
    )
    SELECT value, min(tier) AS tier, max(count) AS cnt
      FROM ranked GROUP BY value ORDER BY tier, cnt DESC, value
    """
    order = [r["value"] for r in q(rank_sql, "rank_m", "кр%", "1", "test-parish", "rank_m", "кр%")]
    check("первым идёт значение из текущего дела", order[0] == "крестьянин", " → ".join(order[:4]))
    check("вторым — значение из прихода", order[1] == "крестьянский сын", " → ".join(order[:4]))
    check("словарные варианты уходят вниз",
          len(order) > 2 and order[-1] not in ("крестьянин", "крестьянский сын"))
    check("«купец» не попал в выдачу по префиксу «кр»", "купец" not in order)

    print("\n7. Формы имён и разбор строки ИОФ")
    n_forms = one("SELECT count(*) FROM name_form")
    check("таблица форм заполнена", n_forms > 12000, f"{n_forms} написаний")
    check("у каждой формы есть имя-владелец",
          one("SELECT count(*) FROM name_form f LEFT JOIN name_dict d ON d.id=f.name_id WHERE d.id IS NULL") == 0)
    # Ключевая тонкость разбора: «Никита» есть и как самостоятельное имя,
    # и как вариант «Аникиты». Приоритет обязан быть у самостоятельного,
    # иначе разбор молча портит имена.
    head = db.execute(
        "SELECT d.name FROM name_form f JOIN name_dict d ON d.id=f.name_id "
        "WHERE f.kind IN ('name','variant') AND f.form_norm='никита' ORDER BY f.priority LIMIT 1"
    ).fetchone()[0]
    check("«Никита» не подменяется «Аникитой»", head == "Никита", f"получилось {head!r}")
    # Отчество должно опознаваться и в старой форме, и в современной.
    for form, expect in [("алексеев", "Алексеевич"), ("алексеевич", "Алексеевич"),
                         ("васильева", "Васильевна")]:
        got = db.execute(
            "SELECT CASE WHEN f.kind LIKE '%_m' THEN d.patr_m ELSE d.patr_f END "
            "FROM name_form f JOIN name_dict d ON d.id=f.name_id "
            "WHERE f.kind LIKE 'patr%' AND f.form_norm=? ORDER BY f.priority LIMIT 1", (form,)).fetchone()
        check(f"отчество «{form}» → «{expect}»", got and got[0] == expect,
              f"получилось {got[0] if got else None!r}")

    print("\n8. Словарь имён на живых примерах")
    check("пол имени Иван определяется",
          one("SELECT gender FROM name_dict WHERE name_norm='иван' LIMIT 1") == "М")
    row = db.execute("SELECT patr_m, patr_f, patr_old_m FROM name_dict WHERE name_norm='михаил' LIMIT 1").fetchone()
    check("отчества от «Михаил»", row and row["patr_m"] and row["patr_f"],
          f"{row['patr_m']} / {row['patr_f']} / старое: {row['patr_old_m']}")
    check("варианты написания сохранены",
          one("SELECT count(*) FROM name_dict WHERE variant IS NOT NULL") > 400)

    print("\n9. Настройки по умолчанию")
    stamp = one("SELECT value FROM setting WHERE key='seed_stamp'")
    check("отпечаток поставки записан", bool(stamp) and len(stamp) == 12, stamp)
    check("оба варианта имени сохраняются",
          one("SELECT value FROM setting WHERE key='keep_original_names'") == "1")
    check("автопополнение справочников включено",
          one("SELECT value FROM setting WHERE key='autoextend_lookups'") == "1")

    db.rollback()
    db.close()

    if xlsm_path and xlsm_path.exists():
        print("\n8. Сверка текстовых справочников с исходным Excel")
        warnings.filterwarnings("ignore")
        import openpyxl
        wb = openpyxl.load_workbook(xlsm_path, data_only=True)
        ws = wb["Имена"]
        n_xl = sum(1 for r in range(4, ws.max_row + 1) if ws.cell(r, 2).value not in (None, ""))
        check("имена совпадают с Excel", n_xl == len(read_csv("name_dict.csv")),
              f"Excel {n_xl}, CSV {len(read_csv('name_dict.csv'))}")
        ws = wb["Справочник"]
        lookups = read_csv("lookup.csv")
        for col, kind, title in [(7, "rank_m", "звания мужские"), (2, "archive", "архивы"),
                                 (20, "death_cause", "причины смерти")]:
            n_xl = len({str(ws.cell(r, col).value).strip() for r in range(4, ws.max_row + 1)
                        if ws.cell(r, col).value not in (None, "")})
            n_csv = sum(1 for r in lookups if r["kind"] == kind)
            check(title, n_xl == n_csv, f"Excel {n_xl}, CSV {n_csv}")
    elif xlsm_path:
        print(f"\n8. Пропущено: не найден {xlsm_path}")

    print(f"\nИтог: успешно {ok_count}, ошибок {fail_count}\n")
    return 1 if fail_count else 0


if __name__ == "__main__":
    raise SystemExit(main())
