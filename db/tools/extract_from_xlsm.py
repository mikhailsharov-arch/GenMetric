#!/usr/bin/env python3
"""
Разовое извлечение справочников из Excel-Индексатора в текстовые файлы.

Зачем отдельный шаг. Исходный файл «Индексатор МК *.xlsm» — работа стороннего
автора (А.К., t.me/LinesOfTime), и в репозиторий он не выкладывается. Поэтому
справочники хранятся в git текстом (db/seed/*.csv): их видно в истории,
изменения читаются построчно, и сборка базы не требует наличия xlsm.

Этот скрипт запускается вручную и только тогда, когда нужно перенести в проект
обновлённую версию справочников. Обычная сборка базы его не вызывает.

Запуск:
    python3 db/tools/extract_from_xlsm.py "путь/к/Индексатор МК 20260806.xlsm"
"""

import csv
import re
import sys
import unicodedata
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")
try:
    import openpyxl
except ImportError:
    sys.exit("Нужен openpyxl:  pip install openpyxl")

SEED_DIR = Path(__file__).resolve().parents[1] / "seed"

# Колонки листа «Справочник»: номер -> (kind, заголовок, автопополнение)
LOOKUP_COLUMNS = {
    2:  ("archive",        "Архивы",                    1),
    3:  ("church",         "Церкви",                    1),
    5:  ("uyezd",          "Уезды",                     1),
    6:  ("guberniya",      "Губернии",                  1),
    7:  ("rank_m",         "Звания мужские",            1),
    8:  ("rank_f",         "Звания женские",            1),
    9:  ("rank_clergy",    "Звания церковнослужителей", 1),
    10: ("confession",     "Вероисповедания",           1),
    16: ("kinship",        "Родственники",              0),
    17: ("surname_end_m",  "Окончания фамилий, М",      0),
    18: ("surname_end_f",  "Окончания фамилий, Ж",      0),
    19: ("marriage_order", "Каким браком",              0),
    20: ("death_cause",    "Причины смерти",            1),
    21: ("np_type",        "Типы населённых пунктов",   1),
}

NAME_COLUMNS = [
    ("gender", 1), ("name", 2), ("variant", 3), ("base_name", 4),
    ("usage_note", 5), ("declension", 6), ("genitive", 7),
    ("patr_old_m", 8), ("patr_old_f", 9), ("patr_m", 10), ("patr_f", 11),
    ("source", 12),
]

PLACE_COLUMNS = [
    ("name", 1), ("np_type", 2), ("guberniya", 3), ("uyezd", 4), ("volost", 5),
    ("short_location", 6), ("full_location", 7), ("familio_url", 8),
]


def norm(value) -> str:
    """Ключ для поиска по префиксу. Должен совпадать с norm() в build_seed.py
    и с normalize() в коде приложения."""
    if value is None:
        return ""
    s = unicodedata.normalize("NFC", str(value)).strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", s)


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def write_csv(path: Path, header, rows) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(header)
        w.writerows(rows)
    return len(rows)


def _utf8_stdout() -> None:
    """Windows запускает Python с кодировкой cp1252, и любой вывод кириллицей
    роняет скрипт с UnicodeEncodeError. Переключаем потоки на UTF-8."""
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except (AttributeError, ValueError):
            pass


def main() -> int:
    _utf8_stdout()
    if len(sys.argv) != 2:
        print(__doc__)
        return 1
    xlsm = Path(sys.argv[1])
    if not xlsm.exists():
        print(f"Не найден файл: {xlsm}")
        return 1

    wb = openpyxl.load_workbook(xlsm, data_only=True)
    total = {}

    # --- имена -------------------------------------------------------------
    ws = wb["Имена"]
    rows = []
    for r in range(4, ws.max_row + 1):
        name = clean(ws.cell(r, 2).value)
        if not name:
            continue
        row = [clean(ws.cell(r, col).value) for _, col in NAME_COLUMNS]
        if row[0] not in ("М", "Ж"):
            row[0] = ""
        rows.append(row)
    total["name_dict.csv"] = write_csv(SEED_DIR / "name_dict.csv",
                                       [n for n, _ in NAME_COLUMNS], rows)

    # --- плоские перечни ---------------------------------------------------
    ws = wb["Справочник"]
    lookup_rows, kind_rows = [], []
    for col, (kind, title, autoextend) in LOOKUP_COLUMNS.items():
        kind_rows.append([kind, title, 1, autoextend])
        seen, order = set(), 0
        for r in range(4, ws.max_row + 1):
            v = clean(ws.cell(r, col).value)
            if not v or v in seen:
                continue
            seen.add(v)
            order += 10
            lookup_rows.append([kind, v, order])
    total["lookup.csv"] = write_csv(SEED_DIR / "lookup.csv",
                                    ["kind", "value", "sort_order"], lookup_rows)
    total["lookup_kind.csv"] = write_csv(SEED_DIR / "lookup_kind.csv",
                                         ["kind", "title", "editable", "autoextend"], kind_rows)

    # --- населённые пункты -------------------------------------------------
    # На листе «НП» присланного файла данных нет: справочник у каждого
    # индексатора свой и накапливается по ходу работы.
    ws = wb["НП"]
    rows, seen = [], set()
    for r in range(3, ws.max_row + 1):
        name = clean(ws.cell(r, 1).value)
        if not name:
            continue
        row = [clean(ws.cell(r, col).value) for _, col in PLACE_COLUMNS]
        key = (row[0], row[1], row[3], row[2])
        if key in seen:
            continue
        seen.add(key)
        rows.append(row)
    total["place.csv"] = write_csv(SEED_DIR / "place.csv",
                                   [n for n, _ in PLACE_COLUMNS], rows)

    print(f"Извлечено из {xlsm.name}:")
    for fname, n in total.items():
        print(f"  {fname:<18} {n:>6} строк")
    print(f"\nФайлы записаны в {SEED_DIR}")
    print("Проверьте изменения через git diff перед коммитом.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
